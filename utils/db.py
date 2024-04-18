#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Time    : 2024/04/16 17:20
# @Author  : An Ju
# @Author  : Allen zh
# @Email   : allenzh@outlook.com
# @File    : db.py

import sqlite3
import os
import pandas as pd
import openpyxl
import utils.tools

class DataBase (object) :
    __tableStat=False
    __connStat=False
    __logger=print

    def __init__ (self,bReNew=True, logger=None) :
        if(logger):self.__logger=logger
        self.dbAddress = 'database/'
        self.table = 'playlists'
        self.T = utils.tools.Tools()
        self.T.mkdir(self.dbAddress)
        if(bReNew):
            self.T.del_file(self.dbAddress)
        if self._connect() == False:
            self.__connStat = False
        else :
            self.__connStat = True
            self.chkTable()
        
    def __del__ (self) :
        if self.__connStat == True :
            self.disConn()
            self.conn=None
            self.cur=None

    #连接数据库，返回连接结果
    def _connect (self) :
        try:
            if(self.__connStat == True):
                self.__del__()
            if not os.path.exists(self.dbAddress) :
                os.makedirs(self.dbAddress)
            self.dbAddress += 'db.sqlite3'
            self.conn = sqlite3.connect(self.dbAddress)
            self.cur = self.conn.cursor()
            return True
        except Exception as e:
            self.__logger(e)
            return False

    #初始化建表
    def __create (self) :
        if self.__connStat == False : return False
        if self.__tableStat == True : return True

        sql = f'''create table {self.table} 
                      (id integer PRIMARY KEY autoincrement, 
                      title nvarchar(30), 
                      tvgroup nvarchar(30),
                      url nvarchar(2048),
                      uniquename nvarchar(30), 
                      delay integer, 
                      speed varchar(20), 
                      videosize varchar(30),
                      format varchar(50),
                      tvorder integer null)'''
        self.cur.execute(sql)
        sql ='''create table tvorders (
                    title nvarchar(30) not null primary key,
                    tvgroup nvarchar(30) null,
                    uniquename nvarchar(30) not null,
                    memo text null,
                    tvorder integer null )
                ''' #uniquename:唯一名,比如CCTV1,对应title可能是CCTV1,CCTV-1,CCTV1 综合频道
        self.cur.execute(sql)
        self.__tableStat=True
        return True
        
    #设置节目清单排序
    #bReNew=True,表示删除原表内容
    def set_tvorders(self,xlsfilename=r'playlists/sortlist.xlsx',bReNew=True):
        if self.__connStat == False : return False

        try:
            listinsheet=openpyxl.load_workbook(xlsfilename)
            datainlist=listinsheet.active #获取excel文件当前表格
            if(bReNew==True): 
                self.cur.execute("delete from tvorders ")
            else:
                rows=self.cur.execute('select count(*) from tvorders where tvorder<9999')
                count=rows.fetchone()[0]
                if (count >0):
                    return count

            data_truck='''INSERT INTO tvorders(title,tvgroup,uniquename,memo,tvorder) 
                        VALUES (?,?,?,?,?)'''
            for row in datainlist.iter_rows(min_row=2,max_col=5,max_row=datainlist.max_row): 
            #使excel各行数据成为迭代器
                cargo=[cell.value for cell in row] #敲黑板！！使每行中单元格成为迭代器
                if(cargo[0] != None):   #title
                    self.cur.execute(data_truck,cargo) #敲黑板！写入一行数据到数据库中表mylist
            self.conn.commit()
            #print("导入节目排序表成功！")
            
            return datainlist.max_row-1
        except Exception as e:
            self.__logger(e)
            return 0
   
    def query (self, sql) :
        '''查询并返回result对象'''
        if self.__connStat == False : return False

        self.cur.execute(sql)
        rows = self.cur.fetchall()

        return rows

    def querypd (self, sql) :
        '''查询并返回pandas dataframe对象'''
        if self.__connStat == False : return False

        return pd.read_sql(sql,self.conn)

    def execute (self, sql) :
        ''' 执行SQL execute，不返回结果集，只返回执行结果True/False（异常）
        '''
        try :
            if self.__connStat == False : return False
            self.cur.execute(sql)
            return True
        except :
            return False

    def insert (self, data):
        ''' 插入数据库表数据
        @data:可以是单条记录，也可以是多条记录list
        '''
        if self.__connStat == False : return False

        if(type(data)==dict):
            recordlist=[]
            recordlist.append(data)
        else:
            recordlist=data

        for record in recordlist:    
            keyList = []
            valList = []
            for k, v in record.items():
                keyList.append(k)
                valList.append(str(v).replace('"','\"').replace("'","''"))

            sql = "insert into " + self.table + " (`" + '`, `'.join(keyList) + "`) values ('" + "', '".join(valList) + "')"
            self.cur.execute(sql)
        self.conn.commit()

    def edit (self, id, data):
        '''#更新update数据库表数据
            @id:记录id,主键
            @data:记录json
        '''
        if self.__connStat == False : return False

        param = ''
        for k, v in data.items():
            param = param + ", `%s` = '%s'" %(k, str(v).replace('"','\"').replace("'","''"))

        param = param[1:]

        sql = "update " + self.table + " set %s WHERE id = %s" % (param, id)
        self.cur.execute(sql)
        self.conn.commit()

    def disConn (self) :
        if self.__connStat == False : return False

        self.cur.close()
        self.conn.close()

    #检查数据库表初始化是否完成，如未完成则自动建表
    def chkTable (self) :
        if self.__connStat == False : return False
        
        ret=False
        sql = "SELECT tbl_name FROM sqlite_master WHERE type='table'"
        self.__tableStat = False

        self.cur.execute(sql)
        values = self.cur.fetchall()

        for x in values:
            if self.table in x :
                self.__tableStat = True
                break

        if self.__tableStat == False :
            ret = self.__create()

        return ret

if __name__ == '__main__':
    db = DataBase()
    #print(db._connect())
    #print(db.chkTable())
    print(db.query('select * from playlists'))
