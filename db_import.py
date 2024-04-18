#!/usr/bin/env python
# -*- coding: utf-8 -*-
# not used,only test
import sqlite3
import openpyxl

def import_tvorders(dbconn=None, xlsfilename=r'playlists/sortlist.xlsx',bDeleteOld=True):
    try:
        conn=dbconn if (dbconn != None) else sqlite3.connect('database\db.sqlite3')
        
        try:
            conn.execute(\
            '''
            create table tvorders (
                tvname nvarchar(30) not null primary key,
                tvgroup nvarchar(30) null,
                memo text null,
                tvorder int null )
                ''')
        except:
            pass
        listinsheet=openpyxl.load_workbook(xlsfilename)
        datainlist=listinsheet.active #获取excel文件当前表格
        if(bDeleteOld==True): conn.execute("delete from tvorders ")
        c=conn.cursor()
        data_truck='''INSERT INTO tvorders(tvname,tvgroup,memo,tvorder) 
                    VALUES (?,?,?,?)'''
        for row in datainlist.iter_rows(min_row=2,max_col=4,max_row=datainlist.max_row): 
        #使excel各行数据成为迭代器
            cargo=[cell.value for cell in row] #敲黑板！！使每行中单元格成为迭代器
            c.execute(data_truck,cargo) #敲黑板！写入一行数据到数据库中表mylist
        conn.commit()
        print("导入节目排序表成功！")
       
        if(dbconn==None): conn.close()
        return datainlist.max_row-1
    except Exception as e:
        print(e)
        return 0


if __name__ == '__main__':
    conn=sqlite3.connect('database\db.sqlite3') 
    if(0<import_tvorders(conn)):
        sql='select * from tvorders o order by o.tvorder'   #'select p.id,p.title,o.tvname,o.tvorder from playlists p left join tvorders o  on p.title=o.tvname order by p.tvgroup,o.tvorder '
        rows=conn.execute(sql)
        for row in rows:
            print(row)
    conn.close()